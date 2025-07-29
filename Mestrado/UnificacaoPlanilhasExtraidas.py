import pandas as pd
from openpyxl import load_workbook

#Unificação das Tabelas obtidas via Webscrapping

# Esse código combina os dados das planilhas "Nivel1", "Nivel2" e "Nivel3 obtidos com base na"
# extração via UIPath de um arquivo Excel da OBMEP para o ano designado, adiciona colunas de nível e ano,
# cria uma nova aba chamada "Combined" com os dados unificados,
# remove as abas originais e exporta o resultado final para um arquivo CSV

arquivo_xlsx = "C:\\Users\\Henrique\\Documents\\UiPath\\Teste\\1OBMEP_BRONZE_2016.xlsx"
arquivo_csv = "C:\\Users\\Henrique\\Documents\\UiPath\\Teste\\1OBMEP_BRONZE_2016.csv"

workbook = load_workbook(filename=arquivo_xlsx)

dfs = []

for folha in ['Nivel1', 'Nivel2', 'Nivel3']:
    df = pd.read_excel(arquivo_xlsx, sheet_name=folha, header=None)
    df.drop(df.columns[0], axis=1, inplace=True)
    nivel_num = int(folha[-1])

    df['nivel'] = nivel_num
    df['ano'] = 2016
    dfs.append(df)

combined_df = pd.concat(dfs, ignore_index=True)

if 'Combined' not in workbook.sheetnames:
    with pd.ExcelWriter(arquivo_xlsx, engine='openpyxl', mode='a') as writer:
        combined_df.to_excel(writer, sheet_name='Combined', index=False, header=False)

for folha in ['Nivel1', 'Nivel2', 'Nivel3']:
    if folha in workbook.sheetnames:
        std = workbook[folha]
        workbook.remove(std)

workbook.save(arquivo_xlsx)

combined_df.to_csv(arquivo_csv, index=False, header=False)

print("Processo concluído!")
